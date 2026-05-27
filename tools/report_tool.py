import os
import base64
import pandas as pd
from datetime import datetime
from PIL import Image
import io
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

def calculate_row_height(row_cells, col_widths, default_height=20, font_size=10):
    """
    Estimates the row height based on cell contents and column widths.
    row_cells: a list of (col_idx, cell_value) where col_idx is 1-based index of the column
    col_widths: dict mapping col_idx to width in characters
    """
    max_lines = 1
    for col_idx, val in row_cells:
        if val is None:
            continue
        val_str = str(val)
        col_w = col_widths.get(col_idx, 10)
        if col_w <= 0:
            col_w = 10
            
        lines = val_str.split('\n')
        total_wrapped_lines = 0
        for line in lines:
            line_len = len(line)
            if line_len == 0:
                total_wrapped_lines += 1
                continue
            effective_w = max(1, int(col_w * 0.85))
            wrapped_count = (line_len + effective_w - 1) // effective_w
            total_wrapped_lines += max(1, wrapped_count)
            
        max_lines = max(max_lines, total_wrapped_lines)
    
    estimated_height = max_lines * (font_size * 1.3) + 12
    return max(default_height, estimated_height)

def save_finding_screenshot(screenshot_b64, url):
    """
    Saves a base64 screenshot to the reports/screenshots directory.
    Returns the relative path to the saved image.
    """
    if not screenshot_b64:
        return None
        
    try:
        # Create reports directory if not exists
        reports_dir = "reports"
        screenshots_dir = os.path.join(reports_dir, "screenshots")
        os.makedirs(screenshots_dir, exist_ok=True)
        
        # Generate filename based on timestamp and url
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        clean_url = "".join([c if c.isalnum() else "_" for c in url[:30]])
        filename = f"finding_{timestamp}_{clean_url}.png"
        filepath = os.path.join(screenshots_dir, filename)
        
        # Decode and save image
        image_data = base64.b64decode(screenshot_b64)
        image = Image.open(io.BytesIO(image_data))
        image.save(filepath)
        
        return filepath
    except Exception as e:
        print(f"⚠️ Error saving finding screenshot: {e}")
        return None

def generate_excel_report(findings, output_path=None, title="TESTING REPORT"):
    """
    Generates a professional Excel report from the findings list.
    """
    try:
        if output_path is None:
            reports_dir = "reports"
            os.makedirs(reports_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Extract domain from title
            domain = "website"
            if title and "TEST " in title:
                d = title.split("TEST ")[-1].strip()
                d = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in d)
                if d:
                    domain = d.strip("_")
            output_path = os.path.join(reports_dir, f"{domain}_{timestamp}.xlsx")
            
        # 1. Prepare data
        processed_findings = []
        if not findings:
            # If no issues, create a success notification row
            processed_findings.append({
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "URL": "N/A",
                "Finding / Bug": "✅ Testing Complete",
                "Detailed Description": "No security or UI issues detected during testing.",
                "Evidence Screenshot": ""
            })
        else:
            for f in findings:
                if not isinstance(f, dict): continue
                processed_findings.append({
                    "Timestamp": f.get("timestamp", ""),
                    "URL": f.get("url", ""),
                    "Finding / Bug": f.get("text", ""),
                    "Detailed Description": f.get("details", "No detailed description."),
                    "Evidence Screenshot": f.get("screenshot", "") 
                })
            
        df = pd.DataFrame(processed_findings)
        
        # 2. Create Excel File
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Security Findings', startrow=2) 
        
        workbook = writer.book
        worksheet = writer.sheets['Security Findings']
        
        # 3. Formats & Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=16, color="000000")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # 4. Add Title (Row 1)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title.upper()
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 35

        # 5. Format Header (Row 3)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            
            # Set column widths
            if col_name == "Timestamp": worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            elif col_name == "URL": worksheet.column_dimensions[get_column_letter(col_idx)].width = 45
            elif col_name == "Finding / Bug": worksheet.column_dimensions[get_column_letter(col_idx)].width = 35
            elif col_name == "Detailed Description": worksheet.column_dimensions[get_column_letter(col_idx)].width = 55
            elif col_name == "Evidence Screenshot": worksheet.column_dimensions[get_column_letter(col_idx)].width = 50

        # 6. Insert data and images
        img_col_idx = 5 
        
        for row_idx, finding in enumerate(processed_findings, start=4):
            # Apply border and alignment for all cells in row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx == 4: # Detailed Description -> Left align
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
            
            # Calculate height needed for text
            row_cells = [
                (1, finding.get("Timestamp", "")),
                (2, finding.get("URL", "")),
                (3, finding.get("Finding / Bug", "")),
                (4, finding.get("Detailed Description", ""))
            ]
            col_widths = {1: 15, 2: 45, 3: 35, 4: 55}
            text_height = calculate_row_height(row_cells, col_widths, default_height=20, font_size=10)
            
            # Image handling
            img_path = finding.get("Evidence Screenshot")
            image_height = 0
            if img_path and os.path.exists(img_path):
                try:
                    img = OpenpyxlImage(img_path)
                    # Resize image
                    orig_w, orig_h = img.width, img.height
                    new_w = 350 
                    new_h = int((new_w / orig_w) * orig_h)
                    img.width = new_w
                    img.height = new_h
                    
                    # Add image to cell
                    cell_ref = f"{get_column_letter(img_col_idx)}{row_idx}"
                    worksheet.add_image(img, cell_ref)
                    
                    # Clear path text in image cell
                    worksheet.cell(row=row_idx, column=img_col_idx).value = ""
                    
                    image_height = new_h * 0.75 + 15
                except Exception as img_err:
                    print(f"⚠️ Could not embed image {img_path}: {img_err}")
                    worksheet.cell(row=row_idx, column=img_col_idx).value = "[Image Error]"
            else:
                 worksheet.cell(row=row_idx, column=img_col_idx).value = "[No Image]"
            
            worksheet.row_dimensions[row_idx].height = max(text_height, image_height)

        writer.close()
        return output_path
    except Exception as e:
        print(f"⚠️ Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None

def cleanup_reports():
    """
    Deletes all temporary screenshots and Excel reports in the reports directory.
    """
    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        return
        
    try:
        import shutil
        # Delete entire screenshots directory
        screenshots_dir = os.path.join(reports_dir, "screenshots")
        if os.path.exists(screenshots_dir):
            shutil.rmtree(screenshots_dir)
            os.makedirs(screenshots_dir)
            
        # Delete old excel files
        for f in os.listdir(reports_dir):
            if f.endswith(".xlsx") or f.endswith(".md") or f.endswith(".txt"):
                os.remove(os.path.join(reports_dir, f))
        print("🧹 Cleaned up temporary report files.")
    except Exception as e:
        print(f"⚠️ Error cleaning up reports: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# NEW: generate_excel_report_v2 — EVN QA 3-Sheet Format
# ─────────────────────────────────────────────────────────────────────────────

_SEVERITY_COLORS = {
    "critical": "FF0000",   # Red
    "major":    "FF6600",   # Orange
    "minor":    "FFCC00",   # Yellow
    "trivial":  "99CC00",   # Light green
    "passed":   "00CC44",   # Green
}

def _apply_header_style(ws, row: int, columns: list, fill_color: str = "4472C4"):
    """Apply bold white header style to a row."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = align


def generate_excel_report_v2(
    summary: dict,
    defect_log: list,
    validation_results: list,
    history: list,
    meta: dict,
    executive_summary: str = "",
    recommendations: list = None,
    output_path: str = None,
) -> str:
    """Generate a professional 3-sheet QA Excel report.

    Sheet 1 — Test Summary:
        Overall status, pass/fail/untested counts, executive summary, metadata.

    Sheet 2 — Defect Log:
        Detailed bug records with severity color coding and screenshots.

    Sheet 3 — Action History:
        Full chronological log of agent actions.
    """
    try:
        if output_path is None:
            reports_dir = "reports"
            os.makedirs(reports_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Extract domain from meta base_url
            base_url = meta.get("base_url", "") if isinstance(meta, dict) else ""
            import urllib.parse
            try:
                parsed_url = urllib.parse.urlparse(base_url)
                domain = parsed_url.netloc or parsed_url.path
                if domain.startswith("www."):
                    domain = domain[4:]
                domain = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in domain)
                domain = domain.strip("_")
            except Exception:
                domain = "website"
            if not domain:
                domain = "website"
            output_path = os.path.join(reports_dir, f"{domain}_{timestamp}.xlsx")

        from openpyxl import Workbook
        wb = Workbook()

        # ── SHEET 1: TEST SUMMARY ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "📊 Test Summary"

        # Title
        ws1.merge_cells("A1:F1")
        title_cell = ws1["A1"]
        title_cell.value = f"EVN QA TEST SUMMARY REPORT  |  {meta.get('generated_at', '')}"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 40

        # Metadata block (rows 3-7)
        meta_rows = [
            ("Target URL", meta.get("base_url", "N/A")),
            ("Test Goal", meta.get("goal", "N/A")),
            ("Brain Model", meta.get("brain_model", "N/A")),
            ("Eval Model", meta.get("eval_model", "N/A")),
            ("Early Stop (Critical Bug)", "YES ⚠️" if meta.get("early_stop") else "NO"),
        ]
        label_font = Font(bold=True)
        for i, (label, value) in enumerate(meta_rows, start=3):
            ws1.cell(row=i, column=1, value=label).font = label_font
            ws1.cell(row=i, column=2, value=value)
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 60

        # Summary scorecard (rows 10-15)
        ws1.cell(row=10, column=1, value="SCORECARD").font = Font(bold=True, size=12)
        scorecard = [
            ("Total Checkpoints Validated", summary.get("total_test_cases", 0)),
            ("✅ Passed", summary.get("passed", 0)),
            ("❌ Failed", summary.get("failed", 0)),
            ("⬜ Untested", summary.get("untested", 0)),
            ("Overall Status", summary.get("overall_status", "N/A")),
        ]
        status_colors = {"PASS": "00CC44", "FAIL": "FF0000", "PARTIAL": "FF6600"}
        for i, (label, value) in enumerate(scorecard, start=11):
            cell_label = ws1.cell(row=i, column=1, value=label)
            cell_label.font = Font(bold=True)
            cell_val = ws1.cell(row=i, column=2, value=value)
            if label == "Overall Status":
                color = status_colors.get(str(value), "FFFFFF")
                cell_val.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell_val.font = Font(bold=True, color="FFFFFF" if value != "PARTIAL" else "000000")

        # Executive Summary
        ws1.cell(row=17, column=1, value="Executive Summary").font = Font(bold=True, size=11)
        exec_cell = ws1.cell(row=18, column=1, value=executive_summary or "No summary generated.")
        exec_cell.alignment = Alignment(wrap_text=True)
        ws1.merge_cells("A18:F18")
        
        # Calculate dynamic height for executive summary (spanning columns A to F: effective width ~120 characters)
        total_w = 120
        lines = 0
        for line in (executive_summary or "No summary generated.").split('\n'):
            lines += max(1, (len(line) + total_w - 1) // total_w)
        ws1.row_dimensions[18].height = max(40, lines * 15 + 15)

        # Recommendations
        if recommendations:
            ws1.cell(row=20, column=1, value="Recommendations").font = Font(bold=True, size=11)
            for i, rec in enumerate(recommendations, start=21):
                ws1.cell(row=i, column=1, value=f"• {rec}")
                ws1.merge_cells(f"A{i}:F{i}")
                
                # Dynamic height for recommendations (effective width ~110 characters)
                rec_w = 110
                lines = 0
                for line in rec.split('\n'):
                    lines += max(1, (len(line) + rec_w - 1) // rec_w)
                ws1.row_dimensions[i].height = max(20, lines * 15 + 8)

        # ── SHEET 2: DEFECT DETAIL ────────────────────────────────────────────────
        ws2 = wb.create_sheet("🐛 Defect Detail")
        defect_cols = ["ID", "Severity", "Title", "URL", "Description", "Checkpoint"]
        ws2.append(defect_cols)
        _apply_header_style(ws2, 1, defect_cols, fill_color="C00000")
        ws2.freeze_panes = "A2"

        col_widths_2 = [10, 12, 35, 40, 60, 30]
        for i, w in enumerate(col_widths_2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # Create BugImages sheet
        ws_img = wb.create_sheet("🖼️ BugImages")
        ws_img.append(["Bug ID", "Screenshot"])
        _apply_header_style(ws_img, 1, ["Bug ID", "Screenshot"], fill_color="1F3864")
        ws_img.column_dimensions["A"].width = 15
        ws_img.column_dimensions["B"].width = 40
        ws_img.freeze_panes = "A2"

        img_row = 2

        for row_idx, defect in enumerate(defect_log or [], start=2):
            severity = str(defect.get("severity", "minor")).lower()
            color = _SEVERITY_COLORS.get(severity, "FFFFFF")
            
            # Format Description
            steps = defect.get("steps_to_reproduce", "")
            expected = defect.get("expected", "")
            actual = defect.get("actual", "")
            
            description = f"Steps to Reproduce:\n{steps}\n\nExpected Result:\n{expected}\n\nActual Result:\n{actual}"
            
            bug_id = defect.get("id", f"BUG-{row_idx-1:03d}")
            
            row_data = [
                bug_id,
                defect.get("severity", "Minor"),
                defect.get("title", ""),
                defect.get("url", ""),
                description,
                defect.get("checkpoint", ""),
            ]
            ws2.append(row_data)
            
            # Color severity cell
            sev_cell = ws2.cell(row=row_idx, column=2)
            sev_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sev_cell.font = Font(bold=True, color="FFFFFF" if severity in ("critical", "major") else "000000")
            
            # Apply borders and wrap
            for col_idx in range(1, len(defect_cols) + 1):
                c = ws2.cell(row=row_idx, column=col_idx)
                c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")
                
            # Dynamic row height for defect detail
            row_cells = [
                (1, bug_id),
                (2, defect.get("severity", "Minor")),
                (3, defect.get("title", "")),
                (4, defect.get("url", "")),
                (5, description),
                (6, defect.get("checkpoint", ""))
            ]
            col_widths = {1: 10, 2: 12, 3: 35, 4: 40, 5: 60, 6: 30}
            ws2.row_dimensions[row_idx].height = calculate_row_height(row_cells, col_widths, default_height=80, font_size=10)
            
            # Handle Screenshot
            # Try to find matching validation result to get screenshot
            screenshot_b64 = None
            for vr in validation_results or []:
                if not vr.get("passed"):
                    # Match by checkpoint or url
                    if vr.get("checkpoint") == defect.get("checkpoint") or vr.get("url") == defect.get("url"):
                        screenshot_b64 = vr.get("screenshot")
                        if screenshot_b64:
                            break
            
            if screenshot_b64:
                try:
                    # Decode base64
                    img_data = base64.b64decode(screenshot_b64)
                    img = Image.open(io.BytesIO(img_data))
                    
                    # Resize to fit cell nicely
                    img.thumbnail((300, 300))
                    
                    # Save to BytesIO to use with openpyxl
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    ox_img = OpenpyxlImage(img_byte_arr)
                    
                    # Add to BugImages sheet
                    ws_img.cell(row=img_row, column=1, value=bug_id).alignment = Alignment(vertical="center", horizontal="center")
                    ws_img.cell(row=img_row, column=1).border = border
                    
                    ws_img.add_image(ox_img, f"B{img_row}")
                    ws_img.row_dimensions[img_row].height = img.height * 0.75 + 20 # Dynamic height for image row
                    ws_img.cell(row=img_row, column=2).border = border
                    
                    img_row += 1
                except Exception as e:
                    print(f"Error adding image for {bug_id}: {e}")

        if not defect_log:
            ws2.append(["—", "—", "✅ No defects found", "—", "—", "—"])
            for col_idx in range(1, len(defect_cols) + 1):
                c = ws2.cell(row=ws2.max_row, column=col_idx)
                c.border = border

        # Also embed validation_results with pass/fail coloring
        ws2.append([])
        ws2.append(["", "", "— VALIDATION CHECKPOINT LOG —"])
        val_header_row = ws2.max_row + 1
        val_cols = ["Timestamp", "Checkpoint", "URL", "Status", "Severity", "Discrepancy", "Evidence"]
        ws2.append(val_cols)
        _apply_header_style(ws2, ws2.max_row, val_cols, fill_color="385723")
        for vr in validation_results or []:
            passed = vr.get("passed", True)
            sev = str(vr.get("severity") or "").lower()
            color = _SEVERITY_COLORS.get(sev, "00CC44" if passed else "FF6600")
            ws2.append([
                vr.get("timestamp", ""),
                vr.get("checkpoint", ""),
                vr.get("url", ""),
                "PASS ✅" if passed else "FAIL ❌",
                vr.get("severity") or "—",
                vr.get("discrepancy", ""),
                vr.get("evidence", ""),
            ])
            row_i = ws2.max_row
            status_cell = ws2.cell(row=row_i, column=4)
            status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            status_cell.font = Font(bold=True, color="FFFFFF")
            
            # Apply borders, alignment, and dynamic height
            for col_idx in range(1, len(val_cols) + 1):
                c = ws2.cell(row=row_i, column=col_idx)
                c.border = border
                c.alignment = Alignment(wrap_text=True, vertical="top")
                
            row_cells = [
                (1, vr.get("timestamp", "")),
                (2, vr.get("checkpoint", "")),
                (3, vr.get("url", "")),
                (4, "PASS ✅" if passed else "FAIL ❌"),
                (5, vr.get("severity") or "—"),
                (6, vr.get("discrepancy", "")),
                (7, vr.get("evidence", ""))
            ]
            col_widths = {1: 10, 2: 12, 3: 35, 4: 40, 5: 60, 6: 30, 7: 20}
            ws2.row_dimensions[row_i].height = calculate_row_height(row_cells, col_widths, default_height=20, font_size=10)

        # ── SHEET 3: ACTION HISTORY ───────────────────────────────────────────
        ws3 = wb.create_sheet("📜 Action History")
        hist_cols = ["#", "Action Log Entry"]
        ws3.append(hist_cols)
        _apply_header_style(ws3, 1, hist_cols, fill_color="404040")
        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 120
        ws3.freeze_panes = "A2"

        for i, entry in enumerate(history or [], start=1):
            ws3.append([i, entry])
            c = ws3.cell(row=i + 1, column=2)
            c.alignment = Alignment(wrap_text=True)
            # Color-code by entry type
            if "❌" in entry or "Error" in entry:
                c.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
            elif "✅" in entry:
                c.fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
            elif "🔬" in entry or "Validator" in entry:
                c.fill = PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid")
            elif "📋" in entry or "Reporter" in entry:
                c.fill = PatternFill(start_color="FFF0E0", end_color="FFF0E0", fill_type="solid")
            
            row_cells = [(1, i), (2, entry)]
            col_widths = {1: 6, 2: 120}
            ws3.row_dimensions[i + 1].height = calculate_row_height(row_cells, col_widths, default_height=20, font_size=10)

        # Save
        wb.save(output_path)
        print(f"📊 EVN QA Report (3 sheets) saved: {output_path}")
        return output_path

    except Exception as e:
        print(f"⚠️ Error generating EVN QA Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None


def translate_log_entry(text):
    if not text:
        return ""
    text = text.replace("--- Result: ", "").replace("---", "").strip()
    
    replacements = {
        "click_element": "Nhấp chuột",
        "type_text": "Nhập văn bản",
        "select_option": "Chọn tùy chọn",
        "hover_element": "Di chuột",
        "navigate_to": "Đi tới trang",
        "verify_ui": "Kiểm tra giao diện",
        "report_issue": "Báo cáo lỗi",
        "finish_task": "Hoàn thành kiểm thử",
        "open_page": "Mở trang",
        "screenshot": "Chụp màn hình",
        "Validator": "Kiểm tra",
        "PASSED": "ĐẠT",
        "FAILED": "LỖI",
        "Scoping": "Lập phạm vi",
        "Reporter": "Tổng hợp báo cáo",
        "Manager": "Phân tích/Điều phối",
        "Action": "Thực hiện",
    }
    
    for eng, vie in replacements.items():
        text = text.replace(eng, vie)
        
    try:
        from deep_translator import GoogleTranslator
        if any(c.isalpha() for c in text):
            return GoogleTranslator(source="auto", target="vi").translate(text)
    except Exception:
        pass
    return text


def generate_excel_report_v3(
    summary: dict,
    defect_log: list,
    validation_results: list,
    history: list,
    meta: dict,
    executive_summary: str = "",
    recommendations: list = None,
    output_path: str = None,
) -> str:
    """Generate a highly professional 5-sheet QA Excel report in Vietnamese for non-QA readers.

    Sheet 1 — 📊 Tổng quan:
        Overview dashboard, KPI cards, executive summary, recommendations, glossary.

    Sheet 2 — 🐛 Danh sách lỗi:
        Detailed bug records with severity colors and descriptions.

    Sheet 3 — 📋 Kịch bản kiểm thử:
        All validation checkpoints and their pass/fail status.

    Sheet 4 — 👣 Nhật ký thao tác:
        Chronological log of agent actions, translated and cleaned.

    Sheet 5 — 🖼️ Hình ảnh lỗi:
        Bug screenshot evidence.
    """
    try:
        import urllib.parse
        import re

        if output_path is None:
            reports_dir = "reports"
            os.makedirs(reports_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Extract domain name from base_url
            base_url = meta.get("base_url", "") if isinstance(meta, dict) else ""
            try:
                parsed_url = urllib.parse.urlparse(base_url)
                domain = parsed_url.netloc or parsed_url.path
                if domain.startswith("www."):
                    domain = domain[4:]
                domain = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in domain)
                domain = domain.strip("_")
            except Exception:
                domain = "website"
            
            if not domain:
                domain = "website"
                
            output_path = os.path.join(reports_dir, f"{domain}_{timestamp}.xlsx")

        from openpyxl import Workbook
        wb = Workbook()

        # Styles definition
        border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1")
        )
        
        border_thick_outer = Border(
            left=Side(style="medium", color="475569"),
            right=Side(style="medium", color="475569"),
            top=Side(style="medium", color="475569"),
            bottom=Side(style="medium", color="475569")
        )

        def _style_range(ws, cell_range, border_style=None, fill=None, font_style=None, alignment=None):
            for row in ws[cell_range]:
                for cell in row:
                    if border_style:
                        cell.border = border_style
                    if fill:
                        cell.fill = fill
                    if font_style:
                        cell.font = font_style
                    if alignment:
                        cell.alignment = alignment

        # ── SHEET 1: TỔNG QUAN ──────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "📊 Tổng quan"
        ws1.views.sheetView[0].showGridLines = True

        # Title Block (Navy Theme)
        ws1.merge_cells("A1:G1")
        title_cell = ws1["A1"]
        title_cell.value = "BÁO CÁO TỔNG HỢP KẾT QUẢ KIỂM THỬ TỰ ĐỘNG"
        title_cell.font = Font(name="Segoe UI", bold=True, size=15, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 45

        # Metadata Table (A3:B8)
        meta_rows = [
            ("URL Mục tiêu", meta.get("base_url", "N/A")),
            ("Mục tiêu kiểm thử", meta.get("goal", "N/A")),
            ("Mô hình Brain", meta.get("brain_model", "N/A")),
            ("Mô hình Eval", meta.get("eval_model", "N/A")),
            ("Thời gian xuất báo cáo", meta.get("generated_at", "")),
            ("Dừng sớm do lỗi nghiêm trọng", "CÓ ⚠️" if meta.get("early_stop") else "KHÔNG"),
        ]
        
        meta_header_fill = PatternFill(start_color="2F3E46", end_color="2F3E46", fill_type="solid")
        meta_header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
        
        ws1.cell(row=3, column=1, value="THÔNG TIN").font = meta_header_font
        ws1.cell(row=3, column=1).fill = meta_header_fill
        ws1.cell(row=3, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=3, column=1).border = border
        
        ws1.cell(row=3, column=2, value="CHI TIẾT").font = meta_header_font
        ws1.cell(row=3, column=2).fill = meta_header_fill
        ws1.cell(row=3, column=2).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=3, column=2).border = border
        ws1.row_dimensions[3].height = 25
        
        for idx, (label, val) in enumerate(meta_rows, start=4):
            ws1.cell(row=idx, column=1, value=label).font = Font(name="Segoe UI", bold=True, size=10, color="334155")
            ws1.cell(row=idx, column=1).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            ws1.cell(row=idx, column=1).border = border
            ws1.cell(row=idx, column=1).alignment = Alignment(vertical="center")
            
            c_val = ws1.cell(row=idx, column=2, value=val)
            c_val.font = Font(name="Segoe UI", size=10)
            c_val.border = border
            c_val.alignment = Alignment(wrap_text=True, vertical="center")
            ws1.row_dimensions[idx].height = 22
            
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 45

        # Scorecard / KPI Cards (D3:G8)
        cards_data = [
            ("D3:D5", "TỔNG KỊCH BẢN", summary.get("total_test_cases", 0), "F1F5F9", "475569"),
            ("E3:E5", "ĐẠT ✅", summary.get("passed", 0), "E6F4EA", "137333"),
            ("F3:F5", "LỖI ❌", summary.get("failed", 0), "FCE8E6", "C5221F"),
        ]
        
        overall = str(summary.get("overall_status", "N/A")).upper()
        if overall in ("PASS", "ĐẠT"):
            overall_label = "ĐẠT ✅"
            overall_fill = "E6F4EA"
            overall_color = "137333"
        elif overall in ("FAIL", "KHÔNG ĐẠT"):
            overall_label = "CÓ LỖI ❌"
            overall_fill = "FCE8E6"
            overall_color = "C5221F"
        else:
            overall_label = "MỘT PHẦN ⚠️"
            overall_fill = "FEF7E0"
            overall_color = "B06000"
            
        cards_data.append(("G3:G5", "KẾT QUẢ CHUNG", overall_label, overall_fill, overall_color))
        
        for cell_range, title_label, val, fill_hex, font_hex in cards_data:
            ws1.merge_cells(cell_range)
            first_cell = ws1[cell_range.split(":")[0]]
            first_cell.value = f"{title_label}\n\n{val}"
            
            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            font = Font(name="Segoe UI", bold=True, size=11, color=font_hex)
            align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            _style_range(ws1, cell_range, border_style=border_thick_outer, fill=fill, font_style=font, alignment=align)

        # Merge untested counts at the bottom of the KPI cards
        ws1.merge_cells("D6:G8")
        untested_cell = ws1["D6"]
        untested_val = summary.get("untested", 0)
        untested_cell.value = f"Kịch bản chưa kiểm tra (được bỏ qua hoặc dừng sớm): {untested_val}"
        
        untested_font = Font(name="Segoe UI", italic=True, size=10, color="5F6368")
        untested_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        untested_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        untested_cell.alignment = untested_align
        untested_cell.font = untested_font
        untested_cell.fill = untested_fill
        
        _style_range(ws1, "D6:G8", border_style=border, fill=untested_fill, font_style=untested_font, alignment=untested_align)

        ws1.column_dimensions["C"].width = 3
        ws1.column_dimensions["D"].width = 18
        ws1.column_dimensions["E"].width = 18
        ws1.column_dimensions["F"].width = 18
        ws1.column_dimensions["G"].width = 20

        # Executive Summary (Row 10)
        ws1.cell(row=10, column=1, value="📝 TÓM TẮT DÀNH CHO BAN GIÁM ĐỐC").font = Font(name="Segoe UI", bold=True, size=12, color="1B365D")
        ws1.row_dimensions[10].height = 25
        
        ws1.merge_cells("A11:G14")
        exec_cell = ws1["A11"]
        exec_cell.value = executive_summary or "Không có tóm tắt tổng quan."
        
        exec_font = Font(name="Segoe UI", size=10, color="1E293B")
        exec_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        exec_align = Alignment(wrap_text=True, vertical="top")
        
        exec_cell.alignment = exec_align
        exec_cell.font = exec_font
        exec_cell.fill = exec_fill
        
        _style_range(ws1, "A11:G14", border_style=border, fill=exec_fill, font_style=exec_font, alignment=exec_align)
        
        # Calculate dynamic height for executive summary (spanning columns A to G: effective width ~140 characters)
        total_w = 140
        lines = 0
        for line in (executive_summary or "Không có tóm tắt tổng quan.").split('\n'):
            lines += max(1, (len(line) + total_w - 1) // total_w)
        total_h = max(80, lines * 15 + 15)
        row_h = total_h / 4
        ws1.row_dimensions[11].height = row_h
        ws1.row_dimensions[12].height = row_h
        ws1.row_dimensions[13].height = row_h
        ws1.row_dimensions[14].height = row_h

        # Recommendations (Row 16)
        ws1.cell(row=16, column=1, value="💡 KHUYẾN NGHỊ & HƯỚNG XỬ LÝ").font = Font(name="Segoe UI", bold=True, size=12, color="1B365D")
        ws1.row_dimensions[16].height = 25
        
        curr_row = 17
        if recommendations:
            for rec in recommendations:
                ws1.cell(row=curr_row, column=1, value="•").font = Font(name="Segoe UI", bold=True, size=12, color="1B365D")
                ws1.cell(row=curr_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                
                rec_cell = ws1.cell(row=curr_row, column=2, value=rec)
                rec_cell.font = Font(name="Segoe UI", size=10, color="1E293B")
                rec_cell.alignment = Alignment(wrap_text=True, vertical="center")
                ws1.merge_cells(start_row=curr_row, start_column=2, end_row=curr_row, end_column=7)
                
                row_border = Border(bottom=Side(style="thin", color="E2E8F0"))
                ws1.cell(row=curr_row, column=1).border = row_border
                for c in range(2, 8):
                    ws1.cell(row=curr_row, column=c).border = row_border
                    
                # Dynamic height for recommendations
                rec_w = 115
                lines = 0
                for line in rec.split('\n'):
                    lines += max(1, (len(line) + rec_w - 1) // rec_w)
                ws1.row_dimensions[curr_row].height = max(24, lines * 15 + 8)
                curr_row += 1
        else:
            ws1.cell(row=curr_row, column=1, value="Không có khuyến nghị cụ thể.")
            ws1.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=7)
            ws1.row_dimensions[curr_row].height = 22
            curr_row += 1

        # Glossary
        curr_row += 1
        ws1.cell(row=curr_row, column=1, value="📘 HƯỚNG DẪN & GIẢI THÍCH THUẬT NGỮ KIỂM THỬ (Dành cho người mới)").font = Font(name="Segoe UI", bold=True, size=12, color="1B365D")
        ws1.row_dimensions[curr_row].height = 25
        
        gl_header_row = curr_row + 1
        ws1.cell(row=gl_header_row, column=1, value="Thuật ngữ").font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
        ws1.cell(row=gl_header_row, column=1).fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        ws1.cell(row=gl_header_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=gl_header_row, column=1).border = border
        
        ws1.cell(row=gl_header_row, column=2, value="Giải thích chi tiết bằng ngôn ngữ phổ thông").font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=10)
        ws1.cell(row=gl_header_row, column=2).fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        ws1.cell(row=gl_header_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws1.cell(row=gl_header_row, column=2).border = border
        ws1.merge_cells(start_row=gl_header_row, start_column=2, end_row=gl_header_row, end_column=7)
        for c in range(3, 8):
            ws1.cell(row=gl_header_row, column=c).border = border
            ws1.cell(row=gl_header_row, column=c).fill = PatternFill(start_color="475569", end_color="475569", fill_type="solid")
        ws1.row_dimensions[gl_header_row].height = 25
        
        glossary_data = [
            ("Đạt (Passed) ✅", "Kịch bản hoạt động hoàn toàn chính xác, đúng như kỳ vọng và không phát sinh bất kỳ lỗi nào."),
            ("Lỗi (Failed) ❌", "Hệ thống gặp sự cố, bị lỗi hiển thị, lỗi chức năng hoặc hoạt động không đúng mô tả nghiệp vụ."),
            ("Chưa kiểm tra (Untested) ⬜", "Kịch bản nằm trong kế hoạch ban đầu nhưng chưa được chạy (do bị hủy hoặc hệ thống đã dừng sớm vì gặp lỗi nghiêm trọng trước đó)."),
            ("Mức độ (Severity)", "Tầm ảnh hưởng của lỗi: Nghiêm trọng (Critical - làm hỏng hệ thống/luồng đăng nhập) > Cao (Major - hỏng tính năng phụ/biểu mẫu) > Thấp (Minor/Trivial - lỗi chính tả/lệch giao diện nhỏ)."),
            ("Kịch bản (Checkpoint)", "Tiêu chí cụ thể cần kiểm tra trên website (ví dụ: 'Đăng nhập thành công', 'Kiểm tra hiển thị nút gửi')."),
            ("Nhật ký thao tác (Steps) 👣", "Trình tự cụ thể mà Robot AI đã nhấp chuột, gõ phím hoặc mở trang web trong quá trình test thực tế.")
        ]
        
        gl_row = gl_header_row
        for term, desc in glossary_data:
            gl_row += 1
            cell_term = ws1.cell(row=gl_row, column=1, value=term)
            cell_term.font = Font(name="Segoe UI", bold=True, size=10, color="1E293B")
            cell_term.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            cell_term.alignment = Alignment(horizontal="center", vertical="center")
            cell_term.border = border
            
            cell_desc = ws1.cell(row=gl_row, column=2, value=desc)
            cell_desc.font = Font(name="Segoe UI", size=10, color="334155")
            cell_desc.alignment = Alignment(wrap_text=True, vertical="center")
            cell_desc.border = border
            
            ws1.merge_cells(start_row=gl_row, start_column=2, end_row=gl_row, end_column=7)
            for c in range(3, 8):
                ws1.cell(row=gl_row, column=c).border = border
                
            ws1.row_dimensions[gl_row].height = 28


        # ── SHEET 2: DANH SÁCH LỖI ──────────────────────────────────────────────
        ws2 = wb.create_sheet("🐛 Danh sách lỗi")
        ws2.views.sheetView[0].showGridLines = True
        
        defect_cols = ["Mã lỗi", "Mức độ nghiêm trọng", "Tiêu đề lỗi", "Đường dẫn (URL)", "Mô tả chi tiết lỗi (Tái hiện & Kết quả)", "Kịch bản liên quan"]
        ws2.append(defect_cols)
        _apply_header_style(ws2, 1, defect_cols, fill_color="991B1B")
        ws2.freeze_panes = "A2"
        
        col_widths_2 = [12, 22, 35, 40, 65, 30]
        for i, w in enumerate(col_widths_2, 1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        _SEVERITY_COLORS = {
            "critical": "DC2626",
            "major":    "EA580C",
            "minor":    "EAB308",
            "trivial":  "3B82F6",
            "passed":   "10B981",
        }

        for row_idx, defect in enumerate(defect_log or [], start=2):
            severity = str(defect.get("severity", "minor")).lower()
            color = _SEVERITY_COLORS.get(severity, "FFFFFF")
            
            steps = defect.get("steps_to_reproduce", "")
            expected = defect.get("expected", "")
            actual = defect.get("actual", "")
            description = f"Các bước tái hiện:\n{steps}\n\nKết quả mong muốn:\n{expected}\n\nKết quả thực tế:\n{actual}"
            bug_id = defect.get("id", f"BUG-{row_idx-1:03d}")
            
            row_data = [
                bug_id,
                defect.get("severity", "Minor").upper(),
                defect.get("title", ""),
                defect.get("url", ""),
                description,
                defect.get("checkpoint", ""),
            ]
            ws2.append(row_data)
            
            # Dynamic row height for defect detail
            row_cells = [
                (1, bug_id),
                (2, defect.get("severity", "Minor").upper()),
                (3, defect.get("title", "")),
                (4, defect.get("url", "")),
                (5, description),
                (6, defect.get("checkpoint", ""))
            ]
            col_widths = {1: 12, 2: 22, 3: 35, 4: 40, 5: 65, 6: 30}
            ws2.row_dimensions[row_idx].height = calculate_row_height(row_cells, col_widths, default_height=80, font_size=10)
            
            sev_cell = ws2.cell(row=row_idx, column=2)
            sev_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sev_cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF" if severity in ("critical", "major") else "000000")
            
            row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if row_idx % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, 7):
                c = ws2.cell(row=row_idx, column=col_idx)
                c.border = border
                if col_idx != 2:
                    if row_fill.fill_type:
                        c.fill = row_fill
                if col_idx in (1, 2):
                    c.alignment = Alignment(horizontal="center", vertical="top")
                else:
                    c.alignment = Alignment(wrap_text=True, vertical="top")

        if not defect_log:
            ws2.append(["—", "—", "✅ Không phát hiện lỗi nào", "—", "Tuyệt vời! Robot AI không phát hiện lỗi UI hay bảo mật nào trên website này.", "—"])
            row_idx = ws2.max_row
            ws2.merge_cells(f"C{row_idx}:F{row_idx}")
            success_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            success_font = Font(name="Segoe UI", bold=True, color="137333")
            for c_idx in range(1, 7):
                cell = ws2.cell(row=row_idx, column=c_idx)
                cell.border = border
                cell.fill = success_fill
                cell.alignment = Alignment(vertical="center", horizontal="center")
            ws2.row_dimensions[row_idx].height = 40


        # ── SHEET 3: KỊCH BẢN KIỂM THỬ ──────────────────────────────────────────────
        ws3 = wb.create_sheet("📋 Kịch bản kiểm thử")
        ws3.views.sheetView[0].showGridLines = True
        
        val_cols = ["STT", "Kịch bản kiểm thử (Checkpoint)", "Đường dẫn kiểm tra (URL)", "Trạng thái", "Mức độ lỗi", "Sai lệch chi tiết (nếu lỗi)", "Chi tiết bằng chứng"]
        ws3.append(val_cols)
        _apply_header_style(ws3, 1, val_cols, fill_color="15803D")
        ws3.freeze_panes = "A2"
        
        col_widths_3 = [6, 35, 40, 22, 15, 45, 30]
        for i, w in enumerate(col_widths_3, 1):
            ws3.column_dimensions[get_column_letter(i)].width = w

        for idx, vr in enumerate(validation_results or [], start=1):
            plan_status = vr.get("_plan_status", "")   # set by reporter when building from task_plan
            passed = vr.get("passed", True)
            sev = vr.get("severity") or "—"

            # Determine display status & colors based on plan_status or passed flag
            if plan_status == "skipped":
                status_text = "BỎ QUA ⏭️"
                status_fill = "E2E8F0"   # cool gray
                status_font_color = "475569"
            elif plan_status == "untested":
                status_text = "CHƯA KIỂM TRA ⬜"
                status_fill = "F1F5F9"   # very light gray
                status_font_color = "64748B"
            elif passed:
                status_text = "ĐẠT ✅"
                status_fill = "E6F4EA"   # green
                status_font_color = "137333"
            else:
                status_text = "LỖI ❌"
                status_fill = "FCE8E6"   # red
                status_font_color = "C5221F"

            row_data = [
                idx,
                vr.get("checkpoint", ""),
                vr.get("url", ""),
                status_text,
                sev,
                vr.get("discrepancy", "") or "—",
                vr.get("evidence", "") or "—"
            ]
            ws3.append(row_data)
            
            row_i = ws3.max_row
            
            # Dynamic row height for validation checkpoints
            row_cells = [
                (1, idx),
                (2, vr.get("checkpoint", "")),
                (3, vr.get("url", "")),
                (4, status_text),
                (5, sev),
                (6, vr.get("discrepancy", "") or "—"),
                (7, vr.get("evidence", "") or "—")
            ]
            col_widths = {1: 6, 2: 35, 3: 40, 4: 20, 5: 15, 6: 45, 7: 30}
            ws3.row_dimensions[row_i].height = calculate_row_height(row_cells, col_widths, default_height=28, font_size=10)
            
            status_cell = ws3.cell(row=row_i, column=4)
            status_cell.fill = PatternFill(start_color=status_fill, end_color=status_fill, fill_type="solid")
            status_cell.font = Font(name="Segoe UI", bold=True, color=status_font_color)
                
            sev_cell = ws3.cell(row=row_i, column=5)
            if not passed and plan_status not in ("skipped", "untested") and sev != "—":
                sev_lower = str(sev).lower()
                sev_color = _SEVERITY_COLORS.get(sev_lower, "FFFFFF")
                sev_cell.fill = PatternFill(start_color=sev_color, end_color=sev_color, fill_type="solid")
                sev_cell.font = Font(name="Segoe UI", bold=True, color="FFFFFF" if sev_lower in ("critical", "major") else "000000")
            
            row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if row_i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, 8):
                cell = ws3.cell(row=row_i, column=col_idx)
                cell.border = border
                if col_idx not in (4, 5):
                    if row_fill.fill_type:
                        cell.fill = row_fill
                if col_idx in (1, 4, 5):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        if not validation_results:
            ws3.append(["—", "Không có kịch bản kiểm thử nào được thực hiện.", "—", "—", "—", "—", "—"])
            row_idx = ws3.max_row
            ws3.merge_cells(f"B{row_idx}:G{row_idx}")
            for c_idx in range(1, 8):
                cell = ws3.cell(row=row_idx, column=c_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center")
            ws3.row_dimensions[row_idx].height = 30


        # ── SHEET 4: NHẬT KÝ THAO TÁC ──────────────────────────────────────────────
        ws4 = wb.create_sheet("👣 Nhật ký thao tác")
        ws4.views.sheetView[0].showGridLines = True
        
        hist_cols = ["STT", "Hành động thực hiện", "Phân loại thao tác", "Trạng thái"]
        ws4.append(hist_cols)
        _apply_header_style(ws4, 1, hist_cols, fill_color="334155")
        ws4.freeze_panes = "A2"
        
        ws4.column_dimensions["A"].width = 6
        ws4.column_dimensions["B"].width = 80
        ws4.column_dimensions["C"].width = 25
        ws4.column_dimensions["D"].width = 18

        step_idx = 1
        for entry in history or []:
            if not isinstance(entry, str):
                continue
                
            category = "Phân tích AI 🧠"
            status = "THÔNG TIN"
            clean_entry = entry
            
            node_prefix_match = re.match(r"^\[(.*?)\]\s*(.*)", entry)
            if node_prefix_match:
                node_name = node_prefix_match.group(1).strip().upper()
                clean_entry = node_prefix_match.group(2).strip()
                
                if "ACTION" in node_name:
                    category = "Robot thao tác 🤖"
                elif "VALIDATOR" in node_name:
                    category = "Kiểm tra chất lượng 🔬"
                elif "SCOPING" in node_name:
                    category = "Lập kế hoạch 🔭"
                elif "REPORTER" in node_name:
                    category = "Tổng hợp báo cáo 📋"
            else:
                if any(kw in entry for kw in ["click_element", "type_text", "select_option", "hover_element", "navigate_to"]):
                    category = "Robot thao tác 🤖"
                elif "Validator" in entry:
                    category = "Kiểm tra chất lượng 🔬"
                elif "Scoping" in entry:
                    category = "Lập kế hoạch 🔭"
            
            if "❌" in clean_entry or "Error" in clean_entry or "FAILED" in clean_entry:
                status = "THẤT BẠI ❌"
            elif "✅" in clean_entry or "PASSED" in clean_entry or "success" in clean_entry.lower():
                status = "THÀNH CÔNG ✅"
            
            clean_entry = clean_entry.replace("✅", "").replace("❌", "").replace("⚠️", "").replace("🔬", "").replace("🤖", "").replace("📋", "").replace("🔭", "").strip()
            translated_action = translate_log_entry(clean_entry)
            
            ws4.append([step_idx, translated_action, category, status])
            row_i = ws4.max_row
            
            # Dynamic row height for Action History
            row_cells = [
                (1, step_idx),
                (2, translated_action),
                (3, category),
                (4, status)
            ]
            col_widths = {1: 6, 2: 80, 3: 25, 4: 18}
            ws4.row_dimensions[row_i].height = calculate_row_height(row_cells, col_widths, default_height=25, font_size=10)
            
            status_cell = ws4.cell(row=row_i, column=4)
            if "THÀNH CÔNG" in status:
                status_cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                status_cell.font = Font(name="Segoe UI", bold=True, color="137333")
            elif "THẤT BẠI" in status:
                status_cell.fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
                status_cell.font = Font(name="Segoe UI", bold=True, color="C5221F")
            else:
                status_cell.fill = PatternFill(start_color="F1F3F4", end_color="F1F3F4", fill_type="solid")
                status_cell.font = Font(name="Segoe UI", bold=True, color="5F6368")
                
            row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") if row_i % 2 == 0 else PatternFill(fill_type=None)
            
            for col_idx in range(1, 5):
                cell = ws4.cell(row=row_i, column=col_idx)
                cell.border = border
                if col_idx != 4:
                    if row_fill.fill_type:
                        cell.fill = row_fill
                if col_idx in (1, 3, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    
            step_idx += 1

        if step_idx == 1:
            ws4.append(["—", "Không có nhật ký thao tác nào.", "—", "—"])
            row_idx = ws4.max_row
            ws4.merge_cells(f"B{row_idx}:D{row_idx}")
            for c_idx in range(1, 5):
                cell = ws4.cell(row=row_idx, column=c_idx)
                cell.border = border
                cell.alignment = Alignment(vertical="center", horizontal="center")
            ws4.row_dimensions[row_idx].height = 30


        # ── SHEET 5: HÌNH ẢNH MINH CHỨNG LỖI ──────────────────────────────────────────────
        ws5 = wb.create_sheet("🖼️ Hình ảnh lỗi")
        ws5.views.sheetView[0].showGridLines = True
        
        ws5.append(["Mã lỗi", "Hình ảnh chụp màn hình thực tế"])
        _apply_header_style(ws5, 1, ["Mã lỗi", "Hình ảnh chụp màn hình thực tế"], fill_color="1E3A8A")
        ws5.column_dimensions["A"].width = 15
        ws5.column_dimensions["B"].width = 55
        ws5.freeze_panes = "A2"
        
        img_row = 2
        for row_idx, defect in enumerate(defect_log or [], start=2):
            bug_id = defect.get("id", f"BUG-{row_idx-1:03d}")
            screenshot_b64 = None
            
            for vr in validation_results or []:
                if not vr.get("passed"):
                    if vr.get("checkpoint") == defect.get("checkpoint") or vr.get("url") == defect.get("url"):
                        screenshot_b64 = vr.get("screenshot")
                        if screenshot_b64:
                            break
                            
            screenshot_path = defect.get("screenshot")
            
            img_added = False
            if screenshot_b64:
                try:
                    img_data = base64.b64decode(screenshot_b64)
                    img = Image.open(io.BytesIO(img_data))
                    img.thumbnail((350, 350))
                    
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    ox_img = OpenpyxlImage(img_byte_arr)
                    
                    ws5.cell(row=img_row, column=1, value=bug_id).alignment = Alignment(vertical="center", horizontal="center")
                    ws5.cell(row=img_row, column=1).font = Font(name="Segoe UI", bold=True, size=11)
                    ws5.cell(row=img_row, column=1).border = border
                    
                    ws5.add_image(ox_img, f"B{img_row}")
                    ws5.row_dimensions[img_row].height = img.height * 0.75 + 20
                    ws5.cell(row=img_row, column=2).border = border
                    
                    img_row += 1
                    img_added = True
                except Exception as e:
                    print(f"Error adding base64 image for {bug_id}: {e}")
                    
            if not img_added and screenshot_path and os.path.exists(screenshot_path):
                try:
                    img = Image.open(screenshot_path)
                    img.thumbnail((350, 350))
                    
                    img_byte_arr = io.BytesIO()
                    img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    ox_img = OpenpyxlImage(img_byte_arr)
                    
                    ws5.cell(row=img_row, column=1, value=bug_id).alignment = Alignment(vertical="center", horizontal="center")
                    ws5.cell(row=img_row, column=1).font = Font(name="Segoe UI", bold=True, size=11)
                    ws5.cell(row=img_row, column=1).border = border
                    
                    ws5.add_image(ox_img, f"B{img_row}")
                    ws5.row_dimensions[img_row].height = img.height * 0.75 + 20
                    ws5.cell(row=img_row, column=2).border = border
                    
                    img_row += 1
                    img_added = True
                except Exception as e:
                    print(f"Error adding screenshot path image for {bug_id}: {e}")
                    
            if not img_added:
                ws5.cell(row=img_row, column=1, value=bug_id).alignment = Alignment(vertical="center", horizontal="center")
                ws5.cell(row=img_row, column=1).font = Font(name="Segoe UI", bold=True, size=11)
                ws5.cell(row=img_row, column=1).border = border
                
                no_img_cell = ws5.cell(row=img_row, column=2, value="[Không có ảnh chụp màn hình]")
                no_img_cell.alignment = Alignment(vertical="center", horizontal="center")
                no_img_cell.font = Font(name="Segoe UI", italic=True, color="64748B")
                no_img_cell.border = border
                ws5.row_dimensions[img_row].height = 35
                img_row += 1

        if img_row == 2:
            ws5.append(["—", "Không có hình ảnh lỗi nào được ghi nhận."])
            ws5.merge_cells("B2:C2")
            ws5.cell(row=2, column=1).border = border
            ws5.cell(row=2, column=2).border = border
            ws5.cell(row=2, column=2).alignment = Alignment(vertical="center", horizontal="center")
            ws5.row_dimensions[2].height = 30

        # Save workbook
        wb.save(output_path)
        print(f"📊 Báo cáo kết quả kiểm thử đã lưu: {output_path}")
        return output_path

    except Exception as e:
        print(f"⚠️ Error generating Vietnamese Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None
