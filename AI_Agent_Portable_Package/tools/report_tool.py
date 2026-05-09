import os
import base64
import pandas as pd
from datetime import datetime
from PIL import Image
import io
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

def save_finding_screenshot(screenshot_b64, url):
    """
    Saves a base64 screenshot to the reports/screenshots directory.
    Returns the relative path to the saved image.
    """
    if not screenshot_b64:
        return None
        
    try:
        # Create reports directory if not exists
        reports_dir = "reports"
        screenshots_dir = os.path.join(reports_dir, "screenshots")
        os.makedirs(screenshots_dir, exist_ok=True)
        
        # Generate filename based on timestamp and url
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        clean_url = "".join([c if c.isalnum() else "_" for c in url[:30]])
        filename = f"finding_{timestamp}_{clean_url}.png"
        filepath = os.path.join(screenshots_dir, filename)
        
        # Decode and save image
        image_data = base64.b64decode(screenshot_b64)
        image = Image.open(io.BytesIO(image_data))
        image.save(filepath)
        
        return filepath
    except Exception as e:
        print(f"⚠️ Error saving finding screenshot: {e}")
        return None

def generate_excel_report(findings, output_path=None, title="TESTING REPORT"):
    """
    Generates a professional Excel report from the findings list.
    """
    try:
        if output_path is None:
            reports_dir = "reports"
            os.makedirs(reports_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(reports_dir, f"security_report_{timestamp}.xlsx")
            
        # 1. Prepare data
        processed_findings = []
        if not findings:
            # If no issues, create a success notification row
            processed_findings.append({
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "URL": "N/A",
                "Finding / Bug": "✅ Testing Complete",
                "Detailed Description": "No security or UI issues detected during testing.",
                "Evidence Screenshot": ""
            })
        else:
            for f in findings:
                if not isinstance(f, dict): continue
                processed_findings.append({
                    "Timestamp": f.get("timestamp", ""),
                    "URL": f.get("url", ""),
                    "Finding / Bug": f.get("text", ""),
                    "Detailed Description": f.get("details", "No detailed description."),
                    "Evidence Screenshot": f.get("screenshot", "") 
                })
            
        df = pd.DataFrame(processed_findings)
        
        # 2. Create Excel File
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        df.to_excel(writer, index=False, sheet_name='Security Findings', startrow=2) 
        
        workbook = writer.book
        worksheet = writer.sheets['Security Findings']
        
        # 3. Formats & Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=16, color="000000")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )

        # 4. Add Title (Row 1)
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title.upper()
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 35

        # 5. Format Header (Row 3)
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            
            # Set column widths
            if col_name == "Timestamp": worksheet.column_dimensions[get_column_letter(col_idx)].width = 15
            elif col_name == "URL": worksheet.column_dimensions[get_column_letter(col_idx)].width = 45
            elif col_name == "Finding / Bug": worksheet.column_dimensions[get_column_letter(col_idx)].width = 35
            elif col_name == "Detailed Description": worksheet.column_dimensions[get_column_letter(col_idx)].width = 55
            elif col_name == "Evidence Screenshot": worksheet.column_dimensions[get_column_letter(col_idx)].width = 50

        # 6. Insert data and images
        img_col_idx = 5 
        
        for row_idx, finding in enumerate(processed_findings, start=4):
            # Apply border and alignment for all cells in row
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx == 4: # Detailed Description -> Left align
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
            
            # Image handling
            img_path = finding.get("Evidence Screenshot")
            if img_path and os.path.exists(img_path):
                try:
                    img = OpenpyxlImage(img_path)
                    # Resize image
                    orig_w, orig_h = img.width, img.height
                    new_w = 350 
                    new_h = int((new_w / orig_w) * orig_h)
                    img.width = new_w
                    img.height = new_h
                    
                    # Add image to cell
                    cell_ref = f"{get_column_letter(img_col_idx)}{row_idx}"
                    worksheet.add_image(img, cell_ref)
                    
                    # Clear path text in image cell
                    worksheet.cell(row=row_idx, column=img_col_idx).value = ""
                    
                    # Adjust row height to fit image
                    worksheet.row_dimensions[row_idx].height = new_h * 0.75 + 15
                except Exception as img_err:
                    print(f"⚠️ Could not embed image {img_path}: {img_err}")
                    worksheet.cell(row=row_idx, column=img_col_idx).value = "[Image Error]"
            else:
                 worksheet.cell(row=row_idx, column=img_col_idx).value = "[No Image]"

        writer.close()
        return output_path
    except Exception as e:
        print(f"⚠️ Error generating Excel report: {e}")
        import traceback
        traceback.print_exc()
        return None

def cleanup_reports():
    """
    Deletes all temporary screenshots and Excel reports in the reports directory.
    """
    reports_dir = "reports"
    if not os.path.exists(reports_dir):
        return
        
    try:
        import shutil
        # Delete entire screenshots directory
        screenshots_dir = os.path.join(reports_dir, "screenshots")
        if os.path.exists(screenshots_dir):
            shutil.rmtree(screenshots_dir)
            os.makedirs(screenshots_dir)
            
        # Delete old excel files
        for f in os.listdir(reports_dir):
            if f.endswith(".xlsx") or f.endswith(".md") or f.endswith(".txt"):
                os.remove(os.path.join(reports_dir, f))
        print("🧹 Cleaned up temporary report files.")
    except Exception as e:
        print(f"⚠️ Error cleaning up reports: {e}")
